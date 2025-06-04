# gera_templates_excel.py

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 1) Definição das listas “oficiais” de Empresa, Unidade e Centro de Custo
EMPRESAS = [
    "KFP PE",
    "KFP AL",
    "KFP BA",
    "KFP CE",
    "KFP RN",
    "KFP PB",
    "KFP PR"
]

UNIDADES = [
    "KFP PE",
    "KFP AL",
    "KFP BA",
    "KFP CE",
    "KFP RN",
    "KFP PB",
    "KFP PR"
]

CENTROS_CUSTO = [
    "Brindes Promocionais - 2.2.04.006",
    "Doação de Equipamentos - 2.2.04.014",
    "Outros"
]


def cria_template_retirada(caminho_saida):
    """
    Gera Template_Retirada.xlsx em 'caminho_saida' com:
      - aba 'Dados' (para preenchimento),
      - aba oculta 'Listas' com colunas de validação (Empresa, Unidade, Centro de Custo),
      - DataValidation aplicada em 'Dados' nas colunas A, C e D.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Cabeçalho da aba “Dados”
    ws["A1"] = "Template de Importação - Retirada"
    colunas = [
        "Empresa",
        "Finalidade",
        "Centro de Custo",
        "Unidade",
        "Nome de quem irá retirar",
        "CPF de quem irá retirar",
        "Prazo Limite",       # AAAA-MM-DD
        "Nome do Produto",
        "Nome Técnico",
        "Qtd",
        "Voltagem",
        "Especificações",
        "Link"
    ]
    for idx, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=2, column=idx)
        cel.value = titulo
        ws.column_dimensions[get_column_letter(idx)].width = max(len(titulo) + 5, 15)

    # Criar aba “Listas” (oculta)
    ws_listas = wb.create_sheet(title="Listas")
    for i, e in enumerate(EMPRESAS, start=1):
        ws_listas.cell(row=i, column=1).value = e
    for i, u in enumerate(UNIDADES, start=1):
        ws_listas.cell(row=i, column=2).value = u
    for i, c in enumerate(CENTROS_CUSTO, start=1):
        ws_listas.cell(row=i, column=3).value = c
    ws_listas.sheet_state = "hidden"

    # DataValidation para “Empresa” (coluna A em “Dados”)
    faixa_empresas = f"'Listas'!$A$1:$A${len(EMPRESAS)}"
    dv_empresa = DataValidation(type="list", formula1=faixa_empresas, showDropDown=True)
    dv_empresa.error = 'Selecione uma das empresas disponíveis.'
    dv_empresa.errorTitle = 'Empresa inválida'
    dv_empresa.prompt = 'Escolha uma empresa da lista'
    dv_empresa.promptTitle = 'Empresa'
    dv_empresa.ranges.add("Dados!$A$3:$A$1000")
    ws.add_data_validation(dv_empresa)

    # DataValidation para “Centro de Custo” (coluna C em “Dados”)
    faixa_custo = f"'Listas'!$C$1:$C${len(CENTROS_CUSTO)}"
    dv_centro = DataValidation(type="list", formula1=faixa_custo, showDropDown=True)
    dv_centro.error = 'Selecione um Centro de Custo válido.'
    dv_centro.errorTitle = 'Centro de Custo inválido'
    dv_centro.prompt = 'Escolha um dos centros de custo da lista'
    dv_centro.promptTitle = 'Centro de Custo'
    dv_centro.ranges.add("Dados!$C$3:$C$1000")
    ws.add_data_validation(dv_centro)

    # DataValidation para “Unidade” (coluna D em “Dados”)
    faixa_unidades = f"'Listas'!$B$1:$B${len(UNIDADES)}"
    dv_unidade = DataValidation(type="list", formula1=faixa_unidades, showDropDown=True)
    dv_unidade.error = 'Selecione uma Unidade válida.'
    dv_unidade.errorTitle = 'Unidade inválida'
    dv_unidade.prompt = 'Escolha uma unidade da lista'
    dv_unidade.promptTitle = 'Unidade'
    dv_unidade.ranges.add("Dados!$D$3:$D$1000")
    ws.add_data_validation(dv_unidade)

    # Ajuste de largura para “Prazo Limite” (coluna G)
    ws.column_dimensions["G"].width = 15

    wb.save(caminho_saida)
    print(f"Template Retirada salvo em: {caminho_saida}")


def cria_template_entrega(caminho_saida):
    """
    Gera Template_Entrega.xlsx em 'caminho_saida' com:
      - aba 'Dados',
      - aba oculta 'Listas' (Empresa, Unidade, Centro de Custo),
      - DataValidation em colunas A, C e D de “Dados”.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Cabeçalho “Entrega” (sem colunas de retirada)
    ws["A1"] = "Template de Importação - Entrega"
    colunas = [
        "Empresa",
        "Finalidade",
        "Centro de Custo",
        "Unidade",
        "Prazo Limite",       # AAAA-MM-DD
        "Nome do Produto",
        "Nome Técnico",
        "Qtd",
        "Voltagem",
        "Especificações",
        "Link"
    ]
    for idx, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=2, column=idx)
        cel.value = titulo
        ws.column_dimensions[get_column_letter(idx)].width = max(len(titulo) + 5, 15)

    # Aba “Listas” oculta
    ws_listas = wb.create_sheet(title="Listas")
    for i, e in enumerate(EMPRESAS, start=1):
        ws_listas.cell(row=i, column=1).value = e
    for i, u in enumerate(UNIDADES, start=1):
        ws_listas.cell(row=i, column=2).value = u
    for i, c in enumerate(CENTROS_CUSTO, start=1):
        ws_listas.cell(row=i, column=3).value = c
    ws_listas.sheet_state = "hidden"

    # DataValidation para “Empresa” (coluna A)
    faixa_empresas = f"'Listas'!$A$1:$A${len(EMPRESAS)}"
    dv_empresa = DataValidation(type="list", formula1=faixa_empresas, showDropDown=True)
    dv_empresa.error = 'Selecione uma das empresas disponíveis.'
    dv_empresa.errorTitle = 'Empresa inválida'
    dv_empresa.prompt = 'Escolha uma empresa da lista'
    dv_empresa.promptTitle = 'Empresa'
    dv_empresa.ranges.add("Dados!$A$3:$A$1000")
    ws.add_data_validation(dv_empresa)

    # DataValidation para “Centro de Custo” (coluna C)
    faixa_custo = f"'Listas'!$C$1:$C${len(CENTROS_CUSTO)}"
    dv_centro = DataValidation(type="list", formula1=faixa_custo, showDropDown=True)
    dv_centro.error = 'Selecione um Centro de Custo válido.'
    dv_centro.errorTitle = 'Centro de Custo inválido'
    dv_centro.prompt = 'Escolha um dos centros de custo da lista'
    dv_centro.promptTitle = 'Centro de Custo'
    dv_centro.ranges.add("Dados!$C$3:$C$1000")
    ws.add_data_validation(dv_centro)

    # DataValidation para “Unidade” (coluna D)
    faixa_unidades = f"'Listas'!$B$1:$B${len(UNIDADES)}"
    dv_unidade = DataValidation(type="list", formula1=faixa_unidades, showDropDown=True)
    dv_unidade.error = 'Selecione uma Unidade válida.'
    dv_unidade.errorTitle = 'Unidade inválida'
    dv_unidade.prompt = 'Escolha uma unidade da lista'
    dv_unidade.promptTitle = 'Unidade'
    dv_unidade.ranges.add("Dados!$D$3:$D$1000")
    ws.add_data_validation(dv_unidade)

    wb.save(caminho_saida)
    print(f"Template Entrega salvo em: {caminho_saida}")


if __name__ == "__main__":
    # AQUI: Aponta para app/static/templates_excel, não para static/ no raiz
    pasta_templates = os.path.join(os.path.dirname(__file__), "app", "static", "templates_excel")
    os.makedirs(pasta_templates, exist_ok=True)

    caminho_retirada = os.path.join(pasta_templates, "Template_Retirada.xlsx")
    caminho_entrega = os.path.join(pasta_templates, "Template_Entrega.xlsx")

    cria_template_retirada(caminho_retirada)
    cria_template_entrega(caminho_entrega)

    print("Geração dos templates concluída.")
